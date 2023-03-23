import math
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import openpyxl
from tkinter import filedialog
from datetime import datetime

# Constants
GR_EPSILON = 0.001
Uk = 1.3
Us = 0.9
fgr = 0.95
r = 21.606299212598 / 2 / 12 / 5280 # Convert tire diameter to miles
tire_diameter = 21.606299212598


class GearRatioCalculator(tk.Frame):

    number1 = 3.3
    number2 = 0.1
    EngineResistanceTpercent = 0.15
    
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid()
        self.create_widgets()
        self.gear_ratios_list = []

    def create_widgets(self):
        # Create labels and entry boxes for horsepower, torque, RPMs, and shift RPM
        hp_label = tk.Label(self, text="Maximum horsepower:")
        hp_label.grid()
        self.hp_entry = tk.Entry(self)
        self.hp_entry.grid()

        tq_label = tk.Label(self, text="Maximum torque:")
        tq_label.grid()
        self.tq_entry = tk.Entry(self)
        self.tq_entry.grid()

        hp_rpm_label = tk.Label(self, text="Maximum horsepower RPM:")
        hp_rpm_label.grid()
        self.hp_rpm_entry = tk.Entry(self)
        self.hp_rpm_entry.grid()

        tq_rpm_label = tk.Label(self, text="Maximum torque RPM:")
        tq_rpm_label.grid()
        self.tq_rpm_entry = tk.Entry(self)
        self.tq_rpm_entry.grid()

        rpm_label = tk.Label(self, text="Maximum RPM:")
        rpm_label.grid()
        self.rpm_entry = tk.Entry(self)
        self.rpm_entry.grid()

        shift_rpm_label = tk.Label(self, text="Shift RPM:")
        shift_rpm_label.grid()
        self.shift_rpm_entry = tk.Entry(self)
        self.shift_rpm_entry.grid()

        # Create frame to hold gear ratio entries
        gear_ratio_frame = tk.Frame(self)
        gear_ratio_frame.grid()

        self.gear_ratio_entries = []
        for i in range(1, 7):
            gear_ratio_entry_label = tk.Label(gear_ratio_frame, text=f"Gear {i}:")
            gear_ratio_entry_label.grid(row=i-1, column=0)
            gear_ratio_entry = tk.Entry(gear_ratio_frame)
            gear_ratio_entry.grid(row=i-1, column=1)
            self.gear_ratio_entries.append(gear_ratio_entry)

        gear_number_label = tk.Label(self, text="Final Drive:")
        gear_number_label.grid()

        self.gear_number_entry = tk.Entry(self)
        self.gear_number_entry.grid() 

        gear_ratio_listbox_label = tk.Label(self, text="Gear Ratio List:")
        gear_ratio_listbox_label.pack()

        self.gear_ratio_listbox = tk.Listbox(self)
        self.gear_ratio_listbox.pack()

        for gear_ratio in self.gear_ratios_list:
            self.gear_ratio_listbox.insert(tk.END, gear_ratio)

        self.gear_ratio_listbox.bind("<<ListboxSelect>>", self.update_selected_gear_ratio)

        self.gear_ratio_listbox_scrollbar = tk.Scrollbar(self, command=gear_ratio_listbox.yview)
        self.gear_ratio_listbox_scrollbar.pack()

        self.gear_ratio_listbox.config(yscrollcommand=gear_ratio_listbox_scrollbar.set)


        # Create button to add gear ratios
        add_ratio_button = tk.Button(self, text="Add Gear Ratio", command=self.add_gear_ratio)
        add_ratio_button.grid()

        # Create button to calculate gear ratios and save to Excel file
        calculate_button = tk.Button(self, text="Calculate Gear Ratios", command=self.calculate_gear_ratios)
        calculate_button.grid()

    def add_gear_ratio(self):
        if self.gear_ratio_entries:
            # Get the gear ratios from the user
            gear_ratios = []
            for entry in self.gear_ratio_entries:
                try:
                    gear_ratio = float(entry.get())
                except ValueError:
                    messagebox.showerror("Error", "Please enter a valid gear ratio")
                    return
                gear_ratios.append(gear_ratio)

            # Get the final drive gear ratio from the user
            try:
                final_drive_ratio = float(self.gear_number_entry.get())
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid final drive gear ratio")
                return

            # Append the gear ratios list with the new gear ratios and final drive gear ratio
            gear_ratios.append(final_drive_ratio)

            # Clear the gear ratio and gear number entries in the GUI
            for entry in self.gear_ratio_entries:
                entry.delete(0, tk.END)
            self.gear_number_entry.delete(0, tk.END)

            # Add the gear ratios to the list of ratios
            self.gear_ratios_list.append(gear_ratios)

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            ws.append([datetime.now(), *gear_ratios])
            wb.save(file_path)

    def calculate_gear_ratios(self):
        # Get the required values from the GUI
        hp = float(self.hp_entry.get())
        tq = float(self.tq_entry.get())
        hp_rpm = float(self.hp_rpm_entry.get())
        tq_rpm = float(self.tq_rpm_entry.get())
        shift_rpm = float(self.shift_rpm_entry.get())

        # Get the tire size and RPM from the user
        tire_size = "195/30/17"
        tire_diameter = 21.606299212598

        # Define target engine torque as a percentage of maximum engine torque
        target_et = (self.number1 * hp + self.number2 * tq) * (1 - self.EngineResistanceTpercent)

        # Read the gear ratios from the CSV file
        file_path = os.path.join(os.environ['userprofile'], 'LOCAL', 'NittoStats.xlsx')
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        ws.append([datetime.now(), *calculated])
        wb.save(file_path)

        # User input
        fastest_et = float(input("Enter the fastest ET for the car: "))

        # Calculate the gear ratios using the gear ratios list and the tire size and RPM
        gear_ratios_calculated = []
        for i in range(len(gear_ratios_list)):
            gear_ratio = gear_ratios_list[i][-1]
            gear_ratio_calculated = gear_ratio * float(self.gear_number_entry.get()) * tire_diameter * math.pi / (5280 * 12 * shift_rpm)
            gear_ratios_calculated.append(gear_ratio_calculated)

        # Calculate gear ratio for known ET
        et = math.inf
        gr = 2.0
        while abs(et - target_et) > GR_EPSILON:
            w = shift_rpm * math.pi / 30.0
            hp *= throttle_percent
            TQflywheel = hp * (5252 / shift_rpm)
            TQwheel = TQflywheel * gr * fgr * 0.85
            WheelThrust = TQwheel / r
            Fnd = max_tq * -1 * number1 - max_tq * -1 * number2
            extrapower = WheelThrust - Fnd
            N = w * 0.4
            gripK = Uk * N
            gripS = Us * N
            gripKhp = Uk * N / 5252
            gripKhpx = Uk * N / 5252 * 550
            WheelThrustApplied = WheelThrust - Fnd
            engineAccel = WheelThrustApplied / (w / 32.1768)
            CarAccel = -2
    

            # Calculate ET
            v0 = 0
            v = 0
            s = 0
            dt = 0.001
            while v < 60:
                t = s / v if v != 0 else 0
                TractionForce = min(WheelThrustApplied, gripS * CarAccel)
                if v == 0:
                    ResistanceForce = EngineResistanceTpercent * TQwheel / r
                else:
                    ResistanceForce = EngineResistanceTpercent * TQwheel / r + (0.5 * 1.225 * v ** 2 * 0.018 * 2) + (0.5 * 1.225 * v ** 2 * 0.65 * 2)
                Force = TractionForce - ResistanceForce
                engineAccel = Force / (w / 32.1768)
                if engineAccel < 0:
                    engineAccel = 0
                v += engineAccel * dt
                if v < 0:
                    v = 0
                s += v * dt
            et = s / (60 * 60)

    def calculate_et(self, gear_ratio, target_et, rpm, hp_rpm, tq_rpm):
        # Calculate engine torque for a given gear ratio
        throttle_percent = 1
        hp_ratio = 1 / (hp_rpm * gear_ratio)
        tq_ratio = 1 / (tq_rpm * gear_ratio)
        et = tq_ratio * gear_ratio * throttle_percent
        while et < target_et:
            throttle_percent += 0.01
            et = tq_ratio * gear_ratio * throttle_percentw
        return et

    def save_gear_ratios(self, file_path):
        # Get the file path from the user
        file_path = os.path.join(os.environ['userprofile'], 'LOCAL', 'NittoStats.xlsx')

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        headers = ["Date", "RT", "ET", "MPH", "Gear 1", "Gear 2", "Gear 3",
                   "Gear 4", "Gear 5", "Gear 6", "Final Drive"]

        # Write the gear ratios to the worksheet
        for i, gear_ratio in enumerate(gear_ratios):
            worksheet.cell(row=i+1, column=1, value=gear_ratio[0])
            worksheet.cell(row=i+1, column=2, value=gear_ratio[1])

        # Save the workbook to the specified file path
        workbook.save(file_path)


if __name__ == "__main__":
    # Create a Tkinter root window
    root = tk.Tk()
    root.title("Gear Ratio Calculator")

    # Create a GearRatioCalculator instance and run the mainloop
    app = GearRatioCalculator(master=root)
    app.mainloop()
